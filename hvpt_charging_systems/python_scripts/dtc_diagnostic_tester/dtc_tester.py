"""
dtc_diagnostic_tester/dtc_tester.py
Automated DTC read/clear/validate tool using UDS (ISO 14229).

Usage:
    python dtc_tester.py --channel PCAN_USBBUS1 --ecu BMS

Features:
    - Read all DTCs with extended data
    - Clear DTCs and verify cleared
    - Parametric DTC test (inject fault → verify DTC sets → clear → verify gone)
    - Excel report output
"""
import can
import isotp
import udsoncan
import argparse
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# DTC code lookup table (extend as needed for your project)
DTC_NAMES = {
    0x0A8000: "P0A80 — Battery System Degraded (SoH)",
    0x0D0001: "P0D00 — HV Isolation Fault",
    0x0A7500: "P0A75 — Battery Contactor Stuck",
    0x0A9000: "P0A90 — Battery Over Temperature",
    0x0AE000: "P0AE0 — Battery Overvoltage",
    0x0AF000: "P0AF0 — Battery Undervoltage",
    0x1A0001: "P1A00 — BMS Communication Timeout",
}

STATUS_BIT_NAMES = {
    0: "testFailed",
    1: "testFailedThisDriveCycle",
    2: "pendingDTC",
    3: "confirmedDTC",
    4: "testNotCompletedSinceLastClear",
    5: "testFailedSinceLastClear",
    6: "testNotCompletedThisDriveCycle",
    7: "warningIndicatorRequested",
}


class DTCTester:
    """Automated DTC diagnostic tool using python-udsoncan."""

    def __init__(self, channel: str, interface: str, tx_id: int, rx_id: int,
                 bitrate: int = 500000):
        self.channel = channel
        self.interface = interface
        self.tx_id = tx_id
        self.rx_id = rx_id
        self.bitrate = bitrate
        self.bus = None
        self.client = None
        self.results = []

    def connect(self) -> bool:
        try:
            self.bus = can.interface.Bus(
                channel=self.channel, bustype=self.interface, bitrate=self.bitrate
            )
            addr = isotp.Address(
                isotp.AddressingMode.Normal_11bits,
                txid=self.tx_id, rxid=self.rx_id
            )
            conn = udsoncan.connections.PythonIsoTpConnection(self.bus, addr)
            conn.open()
            self.client = udsoncan.client.Client(conn, request_timeout=3.0)
            self.client.__enter__()
            print(f"[DTC Tester] Connected to ECU (TX={self.tx_id:#x}, RX={self.rx_id:#x})")
            return True
        except Exception as e:
            print(f"[DTC Tester] Connection failed: {e}")
            return False

    def disconnect(self):
        if self.client:
            self.client.__exit__(None, None, None)
        if self.bus:
            self.bus.shutdown()

    def enter_extended_session(self) -> bool:
        try:
            resp = self.client.change_session(
                udsoncan.services.DiagnosticSessionControl.Session.extendedDiagnosticSession
            )
            return resp.positive
        except Exception as e:
            print(f"Session change failed: {e}")
            return False

    def read_all_dtcs(self, status_mask: int = 0xFF) -> list:
        """Read all DTCs matching status mask."""
        try:
            response = self.client.get_dtc_by_status_mask(status_mask)
            dtcs = []
            if response.positive:
                for dtc in response.dtcs:
                    dtc_id = dtc.id
                    status = dtc.status.get_byte_as_int()
                    active_bits = [STATUS_BIT_NAMES[b] for b in range(8) if status & (1 << b)]
                    dtcs.append({
                        'id': dtc_id,
                        'id_hex': f"0x{dtc_id:06X}",
                        'name': DTC_NAMES.get(dtc_id, f"Unknown DTC {dtc_id:#x}"),
                        'status': f"{status:#04x}",
                        'active_bits': ', '.join(active_bits),
                        'confirmed': bool(status & 0x08),
                        'active': bool(status & 0x01),
                    })
            return dtcs
        except Exception as e:
            print(f"Read DTC failed: {e}")
            return []

    def clear_all_dtcs(self) -> bool:
        """Clear all DTCs (group 0xFFFFFF)."""
        try:
            self.enter_extended_session()
            resp = self.client.clear_dtc(group=0xFFFFFF)
            return resp.positive
        except Exception as e:
            print(f"Clear DTC failed: {e}")
            return False

    def test_dtc_lifecycle(self, dtc_id: int, fault_trigger, fault_clear) -> dict:
        """
        Full DTC lifecycle test:
        1. Clear all DTCs
        2. Trigger fault condition (call fault_trigger())
        3. Wait for DTC to set
        4. Verify DTC present
        5. Clear fault, verify DTC clears after drive cycle
        """
        result = {
            'dtc_id': f"{dtc_id:#08x}",
            'dtc_name': DTC_NAMES.get(dtc_id, 'Unknown'),
            'phase': '',
            'verdict': 'FAIL',
            'details': ''
        }

        # Step 1: Clear
        if not self.clear_all_dtcs():
            result['details'] = "ClearDTC failed"
            return result
        time.sleep(0.5)

        # Step 2: Trigger fault
        print(f"  Triggering fault for DTC {dtc_id:#08x}...")
        fault_trigger()
        time.sleep(2.0)  # Wait for fault to set

        # Step 3: Read DTCs
        dtcs = self.read_all_dtcs(0xFF)
        dtc_found = any(d['id'] == dtc_id for d in dtcs)

        if not dtc_found:
            result['phase'] = 'set_detection'
            result['details'] = f"DTC {dtc_id:#08x} did not set after fault trigger"
            fault_clear()
            return result

        print(f"  DTC {dtc_id:#08x} set correctly.")

        # Step 4: Clear fault and DTC
        fault_clear()
        self.clear_all_dtcs()
        time.sleep(0.5)

        # Step 5: Verify cleared
        dtcs_after = self.read_all_dtcs(0x08)  # confirmed only
        still_present = any(d['id'] == dtc_id for d in dtcs_after)

        if still_present:
            result['phase'] = 'clear_verification'
            result['details'] = f"DTC {dtc_id:#08x} still present after clear"
        else:
            result['verdict'] = 'PASS'
            result['phase'] = 'complete'
            result['details'] = "DTC set and cleared correctly"

        return result

    def generate_report(self, dtcs: list, output_path: str):
        """Generate Excel report of DTC findings."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DTC Report"

        HEADER_FILL = PatternFill(start_color="1A2244", end_color="1A2244", fill_type="solid")
        CONFIRMED_FILL = PatternFill(start_color="CC2233", end_color="CC2233", fill_type="solid")
        PENDING_FILL = PatternFill(start_color="FF8800", end_color="FF8800", fill_type="solid")
        CLEAR_FILL = PatternFill(start_color="006622", end_color="006622", fill_type="solid")
        WHITE = Font(color="FFFFFF", bold=True)

        ws['A1'] = f"DTC Report — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A1'].font = Font(size=13, bold=True)
        ws.merge_cells('A1:F1')

        headers = ['DTC ID', 'DTC Name', 'Status Byte', 'Active Bits',
                   'Confirmed', 'Active Now']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = WHITE

        for row, dtc in enumerate(dtcs, 3):
            ws.cell(row=row, column=1, value=dtc['id_hex'])
            ws.cell(row=row, column=2, value=dtc['name'])
            ws.cell(row=row, column=3, value=dtc['status'])
            ws.cell(row=row, column=4, value=dtc['active_bits'])
            ws.cell(row=row, column=5, value='YES' if dtc['confirmed'] else 'NO')
            ws.cell(row=row, column=6, value='YES' if dtc['active'] else 'NO')

            status_cell = ws.cell(row=row, column=5)
            if dtc['confirmed']:
                status_cell.fill = CONFIRMED_FILL
                status_cell.font = WHITE
            elif dtc['active']:
                status_cell.fill = PENDING_FILL
                status_cell.font = WHITE

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

        wb.save(output_path)
        print(f"[DTC Tester] Report saved: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='DTC Diagnostic Tester')
    parser.add_argument('--channel', default='PCAN_USBBUS1')
    parser.add_argument('--interface', default='pcan')
    parser.add_argument('--bitrate', type=int, default=500000)
    parser.add_argument('--tx', type=lambda x: int(x, 0), default=0x741)
    parser.add_argument('--rx', type=lambda x: int(x, 0), default=0x749)
    parser.add_argument('--report', default='reports/dtc_report.xlsx')
    args = parser.parse_args()

    tester = DTCTester(args.channel, args.interface, args.tx, args.rx, args.bitrate)
    if not tester.connect():
        return

    try:
        print("\n[1] Reading all DTCs...")
        tester.enter_extended_session()
        dtcs = tester.read_all_dtcs(0xFF)
        print(f"    Found {len(dtcs)} DTC(s)")
        for d in dtcs:
            print(f"    {d['id_hex']:12} {d['name'][:40]:40} {d['status']} "
                  f"{'(confirmed)' if d['confirmed'] else ''}")

        import os
        os.makedirs(os.path.dirname(args.report), exist_ok=True)
        tester.generate_report(dtcs, args.report)
    finally:
        tester.disconnect()


if __name__ == '__main__':
    main()
