# adas_framework/utilities/report_generator.py
"""
Test Report Generator — ADAS Enterprise Framework.

Generates:
    1. Allure-compatible JSON fragments (allure serve / allure generate)
    2. HTML single-file summary (no external dependencies)
    3. Excel XLSX workbook with per-feature sheets and colour coding
    4. Grafana/InfluxDB metrics push (optional)

Usage:
    generator = ReportGenerator(cfg.report)
    generator.add_result(result)
    generator.finalize()        # Writes all formats
    generator.open_html()       # Opens browser preview
"""
from __future__ import annotations

import json
import os
import time
import html as _html
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

from core.config import ReportConfig
from core.logger import report_log as log


# ─────────────────────────────────────────────────────────────────────────────
# Data model
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class TestResult:
    test_id:    str
    feature:    str
    status:     str           # "PASS" | "FAIL" | "SKIP" | "ERROR"
    duration_ms: float
    asil:       str           = "QM"
    req_ids:    List[str]     = field(default_factory=list)
    dtc_ids:    List[str]     = field(default_factory=list)
    error_msg:  str           = ""
    timestamp:  float         = field(default_factory=time.time)
    build:      str           = ""
    sw_version: str           = ""
    environment: str          = ""

    @property
    def passed(self) -> bool:
        return self.status == "PASS"

    @property
    def emoji(self) -> str:
        return {"PASS": "✅", "FAIL": "❌", "SKIP": "⚠️", "ERROR": "🔴"}.get(
            self.status, "❓"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Colours
# ─────────────────────────────────────────────────────────────────────────────

_COLOR = {
    "PASS":  "#00C851",
    "FAIL":  "#FF4444",
    "SKIP":  "#FFBB33",
    "ERROR": "#CC0000",
}

_EXCEL_COLOR = {
    "PASS":  "00C851",
    "FAIL":  "FF4444",
    "SKIP":  "FFBB33",
    "ERROR": "CC0000",
}


# ─────────────────────────────────────────────────────────────────────────────
# ReportGenerator
# ─────────────────────────────────────────────────────────────────────────────

class ReportGenerator:
    """Collects test results and outputs multiple report formats."""

    def __init__(self, config: ReportConfig):
        self._cfg     = config
        self._results: List[TestResult] = []
        self._start   = time.time()
        Path(config.output_dir).mkdir(parents=True, exist_ok=True)

    # ── Add result ────────────────────────────────────────────────────────────

    def add_result(self, result: TestResult):
        self._results.append(result)

    def add_from_pytest_report(self, report, feature: str = "UNKNOWN",
                                asil: str = "QM",
                                req_ids: List[str] = None):
        """Ingest a pytest TestReport object."""
        if report.when != "call":
            return
        if report.passed:
            status = "PASS"
        elif report.failed:
            status = "FAIL"
        elif report.skipped:
            status = "SKIP"
        else:
            status = "ERROR"

        error_msg = ""
        if hasattr(report, "longreprtext"):
            error_msg = str(report.longreprtext)[:500]
        elif hasattr(report, "longrepr"):
            error_msg = str(report.longrepr)[:500]

        self.add_result(TestResult(
            test_id    = report.nodeid,
            feature    = feature,
            status     = status,
            duration_ms = getattr(report, "duration", 0.0) * 1000,
            asil       = asil,
            req_ids    = req_ids or [],
            error_msg  = error_msg,
        ))

    # ── Generate all ─────────────────────────────────────────────────────────

    def finalize(self):
        """Write all configured report formats."""
        cfg = self._cfg
        if cfg.html_report:
            path = self._write_html()
            log.info(f"HTML report: {path}")
        if cfg.excel_report and _OPENPYXL:
            path = self._write_excel()
            log.info(f"Excel report: {path}")
        if cfg.allure_results_dir:
            count = self._write_allure()
            log.info(f"Allure results: {count} files → {cfg.allure_results_dir}")
        self._log_summary()

    # ── HTML ──────────────────────────────────────────────────────────────────

    def _write_html(self) -> str:
        results = self._results
        total   = len(results)
        passed  = sum(1 for r in results if r.status == "PASS")
        failed  = sum(1 for r in results if r.status == "FAIL")
        skipped = sum(1 for r in results if r.status == "SKIP")
        errors  = sum(1 for r in results if r.status == "ERROR")
        duration_s = time.time() - self._start

        rows = ""
        for r in results:
            color    = _COLOR.get(r.status, "#888")
            req_str  = ", ".join(r.req_ids) or "-"
            err_clip = _html.escape(r.error_msg[:200]) if r.error_msg else ""
            rows += (
                f"<tr>"
                f"<td style='color:{color}'><b>{r.status}</b></td>"
                f"<td>{_html.escape(r.test_id)}</td>"
                f"<td>{r.feature}</td>"
                f"<td>{r.asil}</td>"
                f"<td>{r.duration_ms:.0f} ms</td>"
                f"<td>{req_str}</td>"
                f"<td style='font-size:11px;color:#999'>{err_clip}</td>"
                f"</tr>\n"
            )

        html_content = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="UTF-8">
<title>ADAS Test Report</title>
<style>
  body {{ font-family: 'Segoe UI', sans-serif; background:#1a1a2e; color:#eee; margin:20px; }}
  h1   {{ color:#00b4d8; }}
  .summary {{ display:flex; gap:20px; margin:20px 0; }}
  .card {{ padding:16px 24px; border-radius:8px; text-align:center; }}
  .pass  {{ background:#005f3c; }} .fail {{ background:#7d0000; }}
  .skip  {{ background:#7a5500; }} .total {{ background:#1f4e79; }}
  table {{ width:100%; border-collapse:collapse; }}
  th    {{ background:#16213e; padding:10px; text-align:left; color:#00b4d8; }}
  td    {{ padding:8px; border-bottom:1px solid #333; font-size:13px; }}
  tr:hover {{ background:#222; }}
</style>
</head><body>
<h1>🚗 ADAS Automated Test Report</h1>
<p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Duration: {duration_s:.1f}s</p>
<div class="summary">
  <div class="card total"><div style="font-size:2em">{total}</div><div>Total</div></div>
  <div class="card pass"><div style="font-size:2em">{passed}</div><div>Passed</div></div>
  <div class="card fail"><div style="font-size:2em">{failed}</div><div>Failed</div></div>
  <div class="card skip"><div style="font-size:2em">{skipped+errors}</div><div>Skip/Err</div></div>
</div>
<table>
<tr><th>Status</th><th>Test</th><th>Feature</th><th>ASIL</th><th>Duration</th><th>Requirements</th><th>Error</th></tr>
{rows}
</table>
</body></html>"""

        path = os.path.join(self._cfg.output_dir, "adas_test_report.html")
        with open(path, "w") as fh:
            fh.write(html_content)
        return path

    # ── Excel ─────────────────────────────────────────────────────────────────

    def _write_excel(self) -> str:
        wb = Workbook()

        # ── Summary sheet ──
        ws = wb.active
        ws.title = "Summary"
        headers = ["Status", "Test ID", "Feature", "ASIL", "Duration (ms)",
                   "Requirements", "Error Message"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="16213E")
            cell.alignment = Alignment(horizontal="center")

        for r in self._results:
            row = [r.status, r.test_id, r.feature, r.asil,
                   round(r.duration_ms, 0), ", ".join(r.req_ids),
                   r.error_msg[:200]]
            ws.append(row)
            fill_color = _EXCEL_COLOR.get(r.status, "888888")
            ws.cell(ws.max_row, 1).fill = PatternFill("solid", fgColor=fill_color)

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

        # ── Per-feature sheets ──
        features: Dict[str, List[TestResult]] = {}
        for r in self._results:
            features.setdefault(r.feature, []).append(r)

        for feature, results in features.items():
            ws2 = wb.create_sheet(title=feature[:31])
            ws2.append(headers)
            for r in results:
                ws2.append([r.status, r.test_id, r.feature, r.asil,
                             round(r.duration_ms, 0), ", ".join(r.req_ids),
                             r.error_msg[:200]])
                fill_color = _EXCEL_COLOR.get(r.status, "888888")
                ws2.cell(ws2.max_row, 1).fill = PatternFill("solid", fgColor=fill_color)

        path = os.path.join(self._cfg.output_dir, "adas_test_report.xlsx")
        wb.save(path)
        return path

    # ── Allure JSON ───────────────────────────────────────────────────────────

    def _write_allure(self) -> int:
        out_dir = Path(self._cfg.allure_results_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        count = 0
        for r in self._results:
            allure_status = {
                "PASS": "passed", "FAIL": "failed",
                "SKIP": "skipped", "ERROR": "broken",
            }.get(r.status, "unknown")

            result_json = {
                "uuid":      f"{r.test_id.replace('/', '_').replace('::', '__')}_{int(r.timestamp)}",
                "name":      r.test_id,
                "status":    allure_status,
                "duration":  int(r.duration_ms),
                "start":     int(r.timestamp * 1000),
                "stop":      int((r.timestamp + r.duration_ms / 1000) * 1000),
                "labels": [
                    {"name": "feature",  "value": r.feature},
                    {"name": "severity", "value": r.asil},
                    {"name": "suite",    "value": r.feature},
                ],
                "links": [
                    {"name": req, "url": f"https://jira.example.com/browse/{req}", "type": "issue"}
                    for req in r.req_ids
                ],
                "statusDetails": {
                    "message": r.error_msg[:500] if r.error_msg else None,
                },
            }
            filename = out_dir / f"{result_json['uuid']}-result.json"
            with open(filename, "w") as fh:
                json.dump(result_json, fh, indent=2)
            count += 1
        return count

    # ── Summary ───────────────────────────────────────────────────────────────

    def _log_summary(self):
        total   = len(self._results)
        passed  = sum(1 for r in self._results if r.status == "PASS")
        failed  = sum(1 for r in self._results if r.status == "FAIL")
        rate    = f"{passed/total:.1%}" if total else "N/A"
        log.info(
            f"Report finalized | total={total} passed={passed} "
            f"failed={failed} pass_rate={rate}"
        )

    def summary(self) -> dict:
        total  = len(self._results)
        passed = sum(1 for r in self._results if r.status == "PASS")
        failed = sum(1 for r in self._results if r.status == "FAIL")
        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "pass_rate": round(passed / total, 4) if total else 0.0,
        }
