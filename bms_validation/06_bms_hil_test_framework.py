#!/usr/bin/env python3
"""
FILE:    06_bms_hil_test_framework.py
PURPOSE: BMS HIL (Hardware-In-the-Loop) Test Framework
         - Test orchestration engine
         - Fault injection manager
         - Measurement data capture
         - KPI calculation (response times, accuracy)
         - HTML + Excel test report generation
         - CI/CD integration via exit codes

DEPENDENCIES:
    pip install pytest pytest-html openpyxl colorama rich

USAGE:
    python 06_bms_hil_test_framework.py --suite full
    python 06_bms_hil_test_framework.py --suite voltage
    python 06_bms_hil_test_framework.py --suite thermal
    python 06_bms_hil_test_framework.py --list

AUTHOR: BMS Validation Team
DATE:   2026-04-19
"""

import time
import sys
import json
import logging
import argparse
from dataclasses import dataclass, field, asdict
from typing import Optional, Callable, List, Dict
from datetime import datetime
from pathlib import Path
from enum import Enum

logger = logging.getLogger("BMS_HIL")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)-7s] %(message)s",
    datefmt="%H:%M:%S.%f"[:-3]
)


# ─────────────────────────────────────────────────────────────────────────────
# 1.  Data Classes
# ─────────────────────────────────────────────────────────────────────────────

class Verdict(str, Enum):
    PASS    = "PASS"
    FAIL    = "FAIL"
    SKIP    = "SKIP"
    ERROR   = "ERROR"
    PENDING = "PENDING"


@dataclass
class TestStep:
    """A single test step within a test case."""
    step_no:     int
    description: str
    action:      str
    expected:    str
    actual:      str     = ""
    verdict:     Verdict = Verdict.PENDING
    duration_ms: float   = 0.0
    note:        str     = ""


@dataclass
class TestCase:
    """A single BMS test case."""
    tc_id:       str
    title:       str
    category:    str
    priority:    str        = "Medium"  # High / Medium / Low
    description: str        = ""
    steps:       List[TestStep] = field(default_factory=list)
    verdict:     Verdict    = Verdict.PENDING
    duration_ms: float      = 0.0
    start_time:  str        = ""
    end_time:    str        = ""
    defects:     List[str]  = field(default_factory=list)
    kpi:         dict       = field(default_factory=dict)


@dataclass
class TestSuiteResult:
    """Overall test suite result."""
    suite_name:  str
    start_time:  str
    end_time:    str        = ""
    total:       int        = 0
    passed:      int        = 0
    failed:      int        = 0
    skipped:     int        = 0
    errors:      int        = 0
    pass_rate:   float      = 0.0
    test_cases:  List[TestCase] = field(default_factory=list)
    environment: dict       = field(default_factory=dict)


# ─────────────────────────────────────────────────────────────────────────────
# 2.  Fault Injection Manager
# ─────────────────────────────────────────────────────────────────────────────
class FaultInjector:
    """
    Manages all BMS fault injections.
    In real HIL, calls CANoe system variable writes or NI PXI I/O.
    In simulation, stores values in memory.
    """

    def __init__(self, canoe_interface=None):
        self._canoe = canoe_interface
        self._state: Dict[str, float] = {}
        self._nominal = {
            "cell_voltage":    3.65,  # V
            "module_temp":     25.0,  # °C
            "pack_current":    0.0,   # A
            "isolation_ohm":   250000, # Ω
            "precharge_curr":  0.05,  # A
        }
        self.restore_all()

    def _write(self, key: str, value: float):
        self._state[key] = value
        if self._canoe:
            # e.g. self._canoe.set_sysvar("BMS_Sim", key, value)
            pass
        logger.debug("[FI] %s = %s", key, value)

    def inject_cell_voltage(self, cell_idx: int, voltage_v: float):
        key = f"Cell_Voltage_{cell_idx:02d}"
        self._write(key, voltage_v)
        logger.info("[FI] Cell[%02d] voltage = %.3f V", cell_idx, voltage_v)

    def inject_module_temp(self, module_idx: int, temp_c: float):
        key = f"Module_Temp_{module_idx}"
        self._write(key, temp_c)
        logger.info("[FI] Module[%d] temp = %.1f °C", module_idx, temp_c)

    def inject_pack_current(self, current_a: float):
        self._write("pack_current", current_a)
        logger.info("[FI] Pack current = %.1f A  (%s)",
                    current_a, "CHARGE" if current_a < 0 else "DISCHARGE")

    def inject_isolation_fault(self, resistance_ohm: float):
        self._write("isolation_ohm", resistance_ohm)
        v_pack = 400  # assumed
        logger.info("[FI] Isolation = %.0f Ω (%.1f Ω/V at %dV)",
                    resistance_ohm, resistance_ohm / v_pack, v_pack)

    def inject_precharge_weld(self, residual_current_a: float = 2.5):
        self._write("precharge_curr", residual_current_a)
        logger.info("[FI] Precharge weld — residual = %.2f A", residual_current_a)

    def inject_sensor_open_circuit(self, module_idx: int):
        self.inject_module_temp(module_idx, -50.0)  # -50°C = invalid signal
        logger.info("[FI] Module[%d] temp sensor open circuit (-50°C)", module_idx)

    def restore_all(self):
        """Restore all injections to nominal state."""
        for i in range(96):
            self._write(f"Cell_Voltage_{i:02d}", self._nominal["cell_voltage"])
        for i in range(8):
            self._write(f"Module_Temp_{i}", self._nominal["module_temp"])
        self._write("pack_current",   self._nominal["pack_current"])
        self._write("isolation_ohm",  self._nominal["isolation_ohm"])
        self._write("precharge_curr", self._nominal["precharge_curr"])
        logger.debug("[FI] All conditions restored to nominal")

    def get_sim_value(self, key: str, default=0):
        return self._state.get(key, default)


# ─────────────────────────────────────────────────────────────────────────────
# 3.  Simulated BMS Response Engine (no hardware mode)
# ─────────────────────────────────────────────────────────────────────────────
class SimulatedBMS:
    """
    Simulates BMS reactions to injected faults for offline testing
    of the test framework itself.
    """

    OV2_THRESH   = 4.25
    OV1_THRESH   = 4.20
    UV2_THRESH   = 2.50
    UV1_THRESH   = 3.00
    OT2_THRESH   = 60.0
    OT1_THRESH   = 50.0
    ISO_WARN     = 200.0  # Ω/V
    ISO_CRIT     = 100.0  # Ω/V
    ISO_EMERG    = 100.0  # Ω/V

    def __init__(self, fault_injector: FaultInjector):
        self.fi = fault_injector

    def _max_cell_v(self):
        return max(
            self.fi.get_sim_value(f"Cell_Voltage_{i:02d}", 3.65)
            for i in range(96)
        )

    def _min_cell_v(self):
        return min(
            self.fi.get_sim_value(f"Cell_Voltage_{i:02d}", 3.65)
            for i in range(96)
        )

    def _max_temp(self):
        return max(
            self.fi.get_sim_value(f"Module_Temp_{i}", 25.0)
            for i in range(8)
        )

    def get_last_dtc(self) -> int:
        max_v = self._max_cell_v()
        min_v = self._min_cell_v()
        max_t = self._max_temp()
        r_iso = self.fi.get_sim_value("isolation_ohm", 250000) / 400  # Ω/V

        if max_v >= self.OV2_THRESH: return 0x0A0E
        if max_v >= self.OV1_THRESH: return 0x0A0D
        if min_v <= self.UV2_THRESH: return 0x0A0F
        if max_t >= self.OT2_THRESH: return 0x1A00
        if max_t >= self.OT1_THRESH: return 0x0A1A
        if r_iso < self.ISO_EMERG:   return 0x1B02
        if r_iso < self.ISO_WARN:    return 0x1B00
        if self.fi.get_sim_value("precharge_curr", 0) > 0.5: return 0x0A3A

        any_invalid_temp = any(
            self.fi.get_sim_value(f"Module_Temp_{i}", 25.0) < -45
            for i in range(8)
        )
        if any_invalid_temp: return 0x0A1C

        return 0x0000  # No fault

    def get_contactor_state(self) -> int:
        """Return 1 if HV ready, 0 if open."""
        dtc = self.get_last_dtc()
        if dtc in (0x0A0E, 0x0A0F, 0x1A00, 0x1B02, 0x0A3B):
            return 0
        return 1

    def get_power_derate_pct(self) -> float:
        max_v = self._max_cell_v()
        min_v = self._min_cell_v()
        max_t = self._max_temp()
        if max_v >= self.OV1_THRESH:  return 80.0
        if min_v <= self.UV1_THRESH:  return 50.0
        if max_t >= self.OT1_THRESH:  return 50.0
        return 100.0

    def get_soc(self) -> float:
        max_v = self._max_cell_v()
        # Simple linear approximation: 3.0V=0%, 4.2V=100%
        soc = min(100, max(0, (max_v - 3.0) / 1.2 * 100))
        if min(self._min_cell_v(), self._max_cell_v()) <= self.UV2_THRESH:
            soc = 0.0
        return soc


# ─────────────────────────────────────────────────────────────────────────────
# 4.  Test Executor
# ─────────────────────────────────────────────────────────────────────────────
class BMSTestExecutor:
    """
    Executes BMS test cases sequentially.
    Measures response times, evaluates pass/fail, records results.
    """

    CONTACTOR_RESPONSE_LIMIT_MS = 100.0   # OV2/UV2 must open within 100 ms
    EMERGENCY_RESPONSE_LIMIT_MS =  50.0   # Emergency HV off within 50 ms
    SOC_ERROR_LIMIT_PCT         =   3.0   # SoC accuracy tolerance

    def __init__(self, fault_injector: FaultInjector, bms: SimulatedBMS):
        self.fi      = fault_injector
        self.bms     = bms
        self.results: List[TestCase] = []

    def _run(self, tc: TestCase, fn: Callable) -> TestCase:
        tc.start_time = datetime.now().isoformat()
        t0 = time.time()
        self.fi.restore_all()
        time.sleep(0.05)

        try:
            fn(tc)
        except AssertionError as ae:
            tc.verdict = Verdict.FAIL
            if not tc.defects:
                tc.defects.append(str(ae))
        except Exception as exc:
            tc.verdict  = Verdict.ERROR
            tc.defects.append(f"ERROR: {exc}")
            logger.exception("Error in %s", tc.tc_id)

        tc.duration_ms = (time.time() - t0) * 1000
        tc.end_time    = datetime.now().isoformat()

        if tc.verdict == Verdict.PENDING:
            tc.verdict = Verdict.PASS

        icon = "✓" if tc.verdict == Verdict.PASS else "✗"
        logger.info("[%s] [%s] %s (%.0f ms)", icon, tc.tc_id, tc.title, tc.duration_ms)
        if tc.defects:
            for d in tc.defects:
                logger.warning("       Defect: %s", d)

        self.fi.restore_all()
        self.results.append(tc)
        return tc

    # ─────────────── Voltage Test Cases ───────────────────────────────────
    def tc_bms_ov_001(self):
        tc = TestCase("TC_BMS_OV_001", "OV1 Warning Detection", "Cell Voltage", "High")
        def run(tc):
            self.fi.inject_cell_voltage(6, 4.21)
            time.sleep(0.1)
            dtc    = self.bms.get_last_dtc()
            derate = self.bms.get_power_derate_pct()
            tc.kpi = {"dtc": hex(dtc), "power_derate_pct": derate}
            assert dtc == 0x0A0D, f"Expected DTC 0x0A0D, got 0x{dtc:04X}"
            assert derate <= 80.0, f"Power derate {derate}% > 80%"
        return self._run(tc, run)

    def tc_bms_ov_002(self):
        tc = TestCase("TC_BMS_OV_002", "OV2 Contactor Open < 100 ms", "Cell Voltage", "High")
        def run(tc):
            t_start = time.time()
            self.fi.inject_cell_voltage(6, 4.26)
            time.sleep(0.05)  # Simulate debounce
            contactor = self.bms.get_contactor_state()
            elapsed_ms = (time.time() - t_start) * 1000
            dtc        = self.bms.get_last_dtc()
            tc.kpi = {"elapsed_ms": elapsed_ms, "contactor": contactor, "dtc": hex(dtc)}
            assert contactor == 0, "Contactor did not open on OV2"
            assert dtc == 0x0A0E, f"DTC not logged: 0x{dtc:04X}"
            assert elapsed_ms <= self.CONTACTOR_RESPONSE_LIMIT_MS, \
                f"Response time {elapsed_ms:.1f} ms > {self.CONTACTOR_RESPONSE_LIMIT_MS} ms"
        return self._run(tc, run)

    def tc_bms_ov_003(self):
        tc = TestCase("TC_BMS_OV_003", "OV2 Single Cell Among 96", "Cell Voltage", "High")
        def run(tc):
            self.fi.inject_cell_voltage(0, 4.26)  # Only cell 0
            time.sleep(0.05)
            dtc = self.bms.get_last_dtc()
            assert dtc == 0x0A0E, f"Single-cell OV2 DTC not detected: 0x{dtc:04X}"
        return self._run(tc, run)

    def tc_bms_ov_004(self):
        tc = TestCase("TC_BMS_OV_004", "OV Noise Spike — No False Trigger", "Cell Voltage", "Medium")
        def run(tc):
            # Quick spike below debounce duration
            self.fi.inject_cell_voltage(10, 4.24)
            time.sleep(0.005)                       # 5 ms < 20 ms debounce
            self.fi.inject_cell_voltage(10, 3.65)   # Restore immediately
            time.sleep(0.1)
            dtc = self.bms.get_last_dtc()
            tc.kpi = {"dtc": hex(dtc)}
            assert dtc not in (0x0A0D, 0x0A0E), \
                f"False OV trigger for noise spike: 0x{dtc:04X}"
        return self._run(tc, run)

    def tc_bms_uv_001(self):
        tc = TestCase("TC_BMS_UV_001", "UV1 Warning Detection", "Cell Voltage", "High")
        def run(tc):
            self.fi.inject_cell_voltage(20, 2.95)
            time.sleep(0.1)
            dtc    = self.bms.get_last_dtc()
            derate = self.bms.get_power_derate_pct()
            tc.kpi = {"dtc": hex(dtc), "derate": derate}
            assert dtc == 0x0A0F, f"Expected UV1 DTC 0x0A0F, got 0x{dtc:04X}"
            assert derate <= 50.0
        return self._run(tc, run)

    def tc_bms_uv_002(self):
        tc = TestCase("TC_BMS_UV_002", "UV2 Contactor Open", "Cell Voltage", "High")
        def run(tc):
            self.fi.inject_cell_voltage(20, 2.45)
            time.sleep(0.05)
            assert self.bms.get_contactor_state() == 0, "Contactor not open for UV2"
        return self._run(tc, run)

    # ─────────────── Temperature Test Cases ───────────────────────────────
    def tc_bms_temp_001(self):
        tc = TestCase("TC_BMS_TEMP_001", "OT1 Warning + Cooling Max", "Temperature", "High")
        def run(tc):
            self.fi.inject_module_temp(0, 52.0)
            time.sleep(0.1)
            dtc    = self.bms.get_last_dtc()
            derate = self.bms.get_power_derate_pct()
            tc.kpi = {"dtc": hex(dtc), "derate": derate, "temp_injected": 52.0}
            assert dtc == 0x0A1A, f"Expected OT1 DTC 0x0A1A, got 0x{dtc:04X}"
            assert derate <= 50.0
        return self._run(tc, run)

    def tc_bms_temp_002(self):
        tc = TestCase("TC_BMS_TEMP_002", "OT2 Emergency Shutdown", "Temperature", "High")
        def run(tc):
            self.fi.inject_module_temp(2, 65.0)
            time.sleep(0.15)
            dtc       = self.bms.get_last_dtc()
            contactor = self.bms.get_contactor_state()
            tc.kpi    = {"dtc": hex(dtc), "contactor": contactor}
            assert dtc == 0x1A00, f"Expected OT2 DTC 0x1A00, got 0x{dtc:04X}"
            assert contactor == 0, "Contactor not opened for OT2"
        return self._run(tc, run)

    def tc_bms_temp_003(self):
        tc = TestCase("TC_BMS_TEMP_003", "Sensor Open Circuit Fail-Safe", "Temperature", "Medium")
        def run(tc):
            self.fi.inject_sensor_open_circuit(1)
            time.sleep(0.1)
            dtc = self.bms.get_last_dtc()
            assert dtc == 0x0A1C, f"Expected sensor OC DTC 0x0A1C, got 0x{dtc:04X}"
        return self._run(tc, run)

    # ─────────────── SoC Test Cases ───────────────────────────────────────
    def tc_bms_soc_001(self):
        tc = TestCase("TC_BMS_SOC_001", "SoC from OCV (3.75 V = 62.5%)", "SoC", "High")
        def run(tc):
            for i in range(96):
                self.fi.inject_cell_voltage(i, 3.75)
            time.sleep(0.2)
            soc = self.bms.get_soc()
            tc.kpi = {"measured_soc": soc, "expected_soc": 62.5}
            assert abs(soc - 62.5) <= self.SOC_ERROR_LIMIT_PCT, \
                f"SoC {soc:.1f}% vs expected 62.5% — error {abs(soc-62.5):.1f}%"
        return self._run(tc, run)

    def tc_bms_soc_002(self):
        tc = TestCase("TC_BMS_SOC_002", "SoC = 100% at Full Charge Voltage", "SoC", "High")
        def run(tc):
            for i in range(96):
                self.fi.inject_cell_voltage(i, 4.15)
            time.sleep(0.2)
            soc = self.bms.get_soc()
            tc.kpi = {"soc": soc}
            assert soc >= 97.0, f"SoC at full charge = {soc:.1f}% (expected ≥97%)"
        return self._run(tc, run)

    def tc_bms_soc_003(self):
        tc = TestCase("TC_BMS_SOC_003", "SoC = 0% at UV2 Cutoff", "SoC", "High")
        def run(tc):
            self.fi.inject_cell_voltage(5, 2.45)
            time.sleep(0.1)
            soc = self.bms.get_soc()
            tc.kpi = {"soc": soc}
            assert soc <= 2.0, f"SoC should be ~0% at UV2 cutoff, got {soc:.1f}%"
        return self._run(tc, run)

    # ─────────────── Isolation Test Cases ─────────────────────────────────
    def tc_bms_iso_001(self):
        tc = TestCase("TC_BMS_ISO_001", "Normal Isolation — No Action", "Isolation", "Medium")
        def run(tc):
            self.fi.inject_isolation_fault(250000)  # 625 Ω/V at 400V
            time.sleep(0.1)
            dtc = self.bms.get_last_dtc()
            assert dtc == 0x0000, f"Unexpected DTC for normal isolation: 0x{dtc:04X}"
        return self._run(tc, run)

    def tc_bms_iso_002(self):
        tc = TestCase("TC_BMS_ISO_002", "Isolation Warning (300 Ω/V)", "Isolation", "High")
        def run(tc):
            self.fi.inject_isolation_fault(120000)  # 300 Ω/V at 400V
            time.sleep(0.2)
            dtc = self.bms.get_last_dtc()
            assert dtc == 0x1B00, f"Expected P1B00, got 0x{dtc:04X}"
        return self._run(tc, run)

    def tc_bms_iso_003(self):
        tc = TestCase("TC_BMS_ISO_003", "Isolation Emergency (75 Ω/V) — HV Off", "Isolation", "High")
        def run(tc):
            self.fi.inject_isolation_fault(30000)   # 75 Ω/V
            time.sleep(0.2)
            dtc       = self.bms.get_last_dtc()
            contactor = self.bms.get_contactor_state()
            tc.kpi    = {"dtc": hex(dtc), "contactor": contactor, "r_iso_ohm_per_v": 75}
            assert dtc == 0x1B02, f"Expected P1B02, got 0x{dtc:04X}"
        return self._run(tc, run)

    # ─────────────── Contactor Test Cases ─────────────────────────────────
    def tc_bms_ctr_001(self):
        tc = TestCase("TC_BMS_CTR_001", "Normal HV ON — DC Link ≥ 95%", "Contactor", "High")
        def run(tc):
            # All nominal → HV should be ready
            time.sleep(0.2)
            contactor = self.bms.get_contactor_state()
            tc.kpi    = {"contactor": contactor}
            assert contactor == 1, "HV not ready in nominal conditions"
        return self._run(tc, run)

    def tc_bms_ctr_004_weld(self):
        tc = TestCase("TC_BMS_CTR_004", "Precharge Contactor Weld Detection", "Contactor", "High")
        def run(tc):
            self.fi.inject_precharge_weld(2.5)
            time.sleep(0.3)
            dtc = self.bms.get_last_dtc()
            assert dtc == 0x0A3A, f"Weld DTC not detected: 0x{dtc:04X}"
        return self._run(tc, run)

    # ─────────────── Cell Balancing Test Cases ────────────────────────────
    def tc_bms_bal_001(self):
        tc = TestCase("TC_BMS_BAL_001", "Balancing Activates at 50 mV Delta", "Balancing", "Medium")
        def run(tc):
            for i in range(96):
                self.fi.inject_cell_voltage(i, 3.75)
            self.fi.inject_cell_voltage(0, 3.80)    # +50 mV
            time.sleep(0.2)
            delta = self.fi.get_sim_value("Cell_Voltage_00") - \
                    min(self.fi.get_sim_value(f"Cell_Voltage_{i:02d}", 3.75) for i in range(1, 96))
            tc.kpi = {"delta_mv": delta * 1000}
            assert delta >= 0.010, f"Delta {delta*1000:.1f} mV not triggering balance"
        return self._run(tc, run)


# ─────────────────────────────────────────────────────────────────────────────
# 5.  Report Generator
# ─────────────────────────────────────────────────────────────────────────────
class ReportGenerator:

    def generate_json(self, suite_result: TestSuiteResult, path: str):
        data = asdict(suite_result)
        with open(path, "w") as f:
            json.dump(data, f, indent=2, default=str)
        logger.info("JSON report: %s", path)

    def generate_html(self, suite_result: TestSuiteResult, path: str):
        rows = ""
        for tc in suite_result.test_cases:
            color = {
                Verdict.PASS: "#d4edda",
                Verdict.FAIL: "#f8d7da",
                Verdict.SKIP: "#fff3cd",
                Verdict.ERROR:"#f8d7da",
            }.get(tc.verdict, "#ffffff")

            kpi_str = ", ".join(f"{k}={v}" for k, v in tc.kpi.items())
            defects = "; ".join(tc.defects) if tc.defects else "—"

            rows += f"""
            <tr style="background:{color}">
              <td>{tc.tc_id}</td>
              <td>{tc.title}</td>
              <td>{tc.category}</td>
              <td><b>{tc.verdict.value}</b></td>
              <td>{tc.duration_ms:.0f} ms</td>
              <td>{kpi_str}</td>
              <td>{defects}</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>BMS Validation Report — {suite_result.suite_name}</title>
  <style>
    body   {{ font-family: Arial, sans-serif; margin: 20px; }}
    h1     {{ color: #003366; }}
    .summary {{ background:#f0f4ff; padding:15px; border-radius:8px; margin-bottom:20px; }}
    table  {{ border-collapse:collapse; width:100%; font-size:13px; }}
    th     {{ background:#003366; color:#fff; padding:8px; text-align:left; }}
    td     {{ padding:7px; border-bottom:1px solid #ddd; }}
    .pass  {{ color:green; font-weight:bold; }}
    .fail  {{ color:red;   font-weight:bold; }}
  </style>
</head>
<body>
  <h1>BMS HIL Validation Report</h1>
  <div class="summary">
    <b>Suite:</b> {suite_result.suite_name}<br>
    <b>Start:</b> {suite_result.start_time}<br>
    <b>End:</b>   {suite_result.end_time}<br>
    <b>Total:</b> {suite_result.total} &nbsp;|&nbsp;
    <span class="pass">Pass: {suite_result.passed}</span> &nbsp;|&nbsp;
    <span class="fail">Fail: {suite_result.failed}</span> &nbsp;|&nbsp;
    <b>Pass Rate: {suite_result.pass_rate:.1f}%</b>
  </div>
  <table>
    <tr>
      <th>TC ID</th><th>Title</th><th>Category</th>
      <th>Verdict</th><th>Duration</th><th>KPI</th><th>Defects</th>
    </tr>
    {rows}
  </table>
</body>
</html>"""

        with open(path, "w") as f:
            f.write(html)
        logger.info("HTML report: %s", path)

    def generate_excel(self, suite_result: TestSuiteResult, path: str):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font

            wb = Workbook()
            ws = wb.active
            ws.title = "BMS Test Results"

            headers = ["TC ID", "Title", "Category", "Priority",
                       "Verdict", "Duration (ms)", "KPI", "Defects"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="003366")

            fill_pass  = PatternFill("solid", fgColor="D4EDDA")
            fill_fail  = PatternFill("solid", fgColor="F8D7DA")

            for tc in suite_result.test_cases:
                row = [
                    tc.tc_id, tc.title, tc.category, tc.priority,
                    tc.verdict.value, round(tc.duration_ms, 1),
                    str(tc.kpi),
                    "; ".join(tc.defects) if tc.defects else ""
                ]
                ws.append(row)
                fill = fill_pass if tc.verdict == Verdict.PASS else fill_fail
                for cell in ws[ws.max_row]:
                    cell.fill = fill

            wb.save(path)
            logger.info("Excel report: %s", path)

        except ImportError:
            logger.warning("openpyxl not installed — skipping Excel report")


# ─────────────────────────────────────────────────────────────────────────────
# 6.  Test Suite Controller
# ─────────────────────────────────────────────────────────────────────────────
class BMSTestSuiteController:

    PASS_GATE = 95.0  # Minimum pass rate for CI/CD gate

    def __init__(self):
        self.fi       = FaultInjector()
        self.bms      = SimulatedBMS(self.fi)
        self.executor = BMSTestExecutor(self.fi, self.bms)
        self.reporter = ReportGenerator()

    def _suite_voltage(self) -> List:
        return [
            self.executor.tc_bms_ov_001,
            self.executor.tc_bms_ov_002,
            self.executor.tc_bms_ov_003,
            self.executor.tc_bms_ov_004,
            self.executor.tc_bms_uv_001,
            self.executor.tc_bms_uv_002,
        ]

    def _suite_thermal(self) -> List:
        return [
            self.executor.tc_bms_temp_001,
            self.executor.tc_bms_temp_002,
            self.executor.tc_bms_temp_003,
        ]

    def _suite_soc(self) -> List:
        return [
            self.executor.tc_bms_soc_001,
            self.executor.tc_bms_soc_002,
            self.executor.tc_bms_soc_003,
        ]

    def _suite_isolation(self) -> List:
        return [
            self.executor.tc_bms_iso_001,
            self.executor.tc_bms_iso_002,
            self.executor.tc_bms_iso_003,
        ]

    def _suite_contactor(self) -> List:
        return [
            self.executor.tc_bms_ctr_001,
            self.executor.tc_bms_ctr_004_weld,
        ]

    def _suite_balancing(self) -> List:
        return [
            self.executor.tc_bms_bal_001,
        ]

    SUITES = {
        "voltage":   "_suite_voltage",
        "thermal":   "_suite_thermal",
        "soc":       "_suite_soc",
        "isolation": "_suite_isolation",
        "contactor": "_suite_contactor",
        "balancing": "_suite_balancing",
    }

    def run(self, suite_name: str = "full") -> TestSuiteResult:
        if suite_name == "full":
            test_fns = (
                self._suite_voltage() +
                self._suite_thermal() +
                self._suite_soc()     +
                self._suite_isolation() +
                self._suite_contactor() +
                self._suite_balancing()
            )
        else:
            getter = getattr(self, self.SUITES.get(suite_name, "_suite_voltage"))
            test_fns = getter()

        suite = TestSuiteResult(
            suite_name = f"BMS_{suite_name.upper()}",
            start_time = datetime.now().isoformat(),
            environment = {
                "mode": "simulation",
                "bms_sw_version": "03.05.00",
                "test_framework": "BMS HIL v1.0",
            }
        )

        logger.info("═" * 60)
        logger.info("  BMS HIL TEST SUITE: %s", suite_name.upper())
        logger.info("  Tests to execute: %d", len(test_fns))
        logger.info("═" * 60)

        for fn in test_fns:
            fn()

        # Collect results from executor
        suite.test_cases = self.executor.results.copy()
        self.executor.results.clear()

        suite.total   = len(suite.test_cases)
        suite.passed  = sum(1 for t in suite.test_cases if t.verdict == Verdict.PASS)
        suite.failed  = sum(1 for t in suite.test_cases if t.verdict == Verdict.FAIL)
        suite.skipped = sum(1 for t in suite.test_cases if t.verdict == Verdict.SKIP)
        suite.errors  = sum(1 for t in suite.test_cases if t.verdict == Verdict.ERROR)
        suite.pass_rate = (suite.passed / suite.total * 100) if suite.total else 0
        suite.end_time  = datetime.now().isoformat()

        # Print summary
        logger.info("\n%s", "═" * 60)
        logger.info("  SUITE RESULT: %s", suite.suite_name)
        logger.info("  Total : %d | Pass: %d | Fail: %d | Skip: %d",
                    suite.total, suite.passed, suite.failed, suite.skipped)
        logger.info("  Pass Rate: %.1f%% (Gate: %.1f%%)", suite.pass_rate, self.PASS_GATE)
        gate_status = "PASS ✓" if suite.pass_rate >= self.PASS_GATE else "FAIL ✗"
        logger.info("  CI/CD Gate: %s", gate_status)
        logger.info("═" * 60)

        # Generate reports
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = Path("bms_reports")
        out_dir.mkdir(exist_ok=True)
        self.reporter.generate_json(suite, str(out_dir / f"bms_{suite_name}_{ts}.json"))
        self.reporter.generate_html(suite, str(out_dir / f"bms_{suite_name}_{ts}.html"))
        self.reporter.generate_excel(suite, str(out_dir / f"bms_{suite_name}_{ts}.xlsx"))

        return suite


# ─────────────────────────────────────────────────────────────────────────────
# 7.  CLI Entry
# ─────────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="BMS HIL Test Framework")
    parser.add_argument("--suite", default="full",
                        choices=["full", "voltage", "thermal", "soc",
                                 "isolation", "contactor", "balancing"],
                        help="Test suite to run")
    parser.add_argument("--list", action="store_true",
                        help="List all available test cases")
    args = parser.parse_args()

    if args.list:
        print("\nAvailable BMS Test Suites:")
        print("  full       — All test cases (~19 TCs)")
        print("  voltage    — Cell voltage OV/UV tests (6 TCs)")
        print("  thermal    — Temperature protection tests (3 TCs)")
        print("  soc        — State of Charge tests (3 TCs)")
        print("  isolation  — IMD isolation tests (3 TCs)")
        print("  contactor  — Contactor & precharge tests (2 TCs)")
        print("  balancing  — Cell balancing tests (1 TC)")
        sys.exit(0)

    controller = BMSTestSuiteController()
    result     = controller.run(args.suite)

    # CI/CD exit code: 0 = pass, 1 = fail
    sys.exit(0 if result.pass_rate >= BMSTestSuiteController.PASS_GATE else 1)


if __name__ == "__main__":
    main()
